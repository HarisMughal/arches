from datetime import datetime
import json
import logging
import math
import os
import uuid
import zipfile
import arches.app.tasks as tasks
from django.core.files import File
from django.http import HttpResponse
from openpyxl import load_workbook
from django.db import connection
from django.utils.translation import ugettext as _
from django.core.files.storage import default_storage
from arches.app.datatypes.datatypes import DataTypeFactory
from arches.app.models.models import Node, NodeGroup
from arches.app.utils.betterJSONSerializer import JSONSerializer
from arches.app.utils.file_validator import FileValidator
from arches.app.utils.index_database import index_resources_by_transaction
from arches.management.commands.etl_template import create_workbook
from openpyxl.writer.excel import save_virtual_workbook
from arches.app.etl_modules.base_import_module import BaseImportModule
from arches.app.etl_modules.branch_excel_importer import BranchExcelImporter

logger = logging.getLogger(__name__)


class TileExcelImporter(BranchExcelImporter):
    def __init__(self, request=None, loadid=None, temp_dir=None):
        self.request = request if request else None
        self.userid = request.user.id if request else None
        self.moduleid = request.POST.get("module") if request else None
        self.datatype_factory = DataTypeFactory()
        self.legacyid_lookup = {}
        self.temp_path = ""
        self.loadid = loadid if loadid else None
        self.temp_dir = temp_dir if temp_dir else None


    def create_tile_value(self, cell_values, data_node_lookup, node_lookup, nodegroup_alias, row_details, cursor):
        node_value_keys = data_node_lookup[nodegroup_alias]
        tile_value = {}
        tile_valid = True
        for key in node_value_keys:
            try:
                nodeid = node_lookup[key]["nodeid"]
                node_details = node_lookup[key]
                datatype = node_details["datatype"]
                datatype_instance = self.datatype_factory.get_instance(datatype)
                source_value = row_details[key]
                config = node_details["config"]
                config["path"] = os.path.join("uploadedfiles", "tmp", self.loadid)
                config["loadid"] = self.loadid
                try:
                    config["nodeid"] = nodeid
                except TypeError:
                    config = {}

                value, validation_errors = self.prepare_data_for_loading(datatype_instance, source_value, config)
                valid = True if len(validation_errors) == 0 else False
                if not valid:
                    tile_valid = False
                error_message = ""
                for error in validation_errors:
                    error_message = "{0}|{1}".format(error_message, error["message"]) if error_message != "" else error["message"]
                    cursor.execute(
                        """
                        INSERT INTO load_errors (type, value, source, error, message, datatype, loadid, nodeid)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                        ("node", source_value, "", error["title"], error["message"], datatype, self.loadid, nodeid),
                    )

                tile_value[nodeid] = {"value": value, "valid": valid, "source": source_value, "notes": error_message, "datatype": datatype}
            except KeyError:
                pass

        tile_value_json = JSONSerializer().serialize(tile_value)
        return tile_value_json, tile_valid

    def process_worksheet(self, worksheet, cursor, node_lookup, nodegroup_lookup):
        data_node_lookup = {}
        nodegroup_tile_lookup = {}
        previous_tile = {}
        row_count = 0
        

        nodegroupid_column = int(worksheet.max_column)
        nodegroup_alias = nodegroup_lookup[worksheet.cell(row=2,column=nodegroupid_column).value]['alias']
        data_node_lookup[nodegroup_alias] = [val.value for val in worksheet[1][3:-3]]

        
        for row in worksheet.iter_rows(min_row=2):
            cell_values = [cell.value for cell in row]
            if len(cell_values) == 0 or any(cell_values) is False:
                continue
            resourceid = cell_values[2]
            if resourceid is None:
                cursor.execute(
                    """UPDATE load_event SET status = %s, load_end_time = %s WHERE loadid = %s""",
                    ("failed", datetime.now(), self.loadid),
                )
                raise ValueError(_("All rows must have a valid resource id"))

            node_values = cell_values[3:-3]
            try:
                row_count += 1
                row_details = dict(zip(data_node_lookup[nodegroup_alias], node_values))
                row_details["nodegroup_id"] = node_lookup[nodegroup_alias]["nodeid"]
                tileid = cell_values[0]
                nodegroup_depth = nodegroup_lookup[row_details["nodegroup_id"]]["depth"]
                parenttileid = None if "None" else cell_values[1]
                parenttileid = self.get_parent_tileid(
                    nodegroup_depth, str(tileid), previous_tile, nodegroup_alias, nodegroup_tile_lookup
                )
                legacyid, resourceid = self.set_legacy_id(resourceid)
                tile_value_json, passes_validation = self.create_tile_value(
                    cell_values, data_node_lookup, node_lookup, nodegroup_alias, row_details, cursor
                )
                operation = 'insert'
                cursor.execute(
                    """INSERT INTO load_staging (nodegroupid, legacyid, resourceid, tileid, parenttileid, value, loadid, nodegroup_depth, source_description, passes_validation, operation) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                    (
                        row_details["nodegroup_id"],
                        legacyid,
                        resourceid,
                        tileid,
                        parenttileid,
                        tile_value_json,
                        self.loadid,
                        nodegroup_depth,
                        "worksheet:{0}, row:{1}".format(worksheet.title, row[0].row),  # source_description
                        passes_validation,
                        operation,
                    ),
                )
            except KeyError:
                pass
        cursor.execute("""CALL __arches_check_tile_cardinality_violation_for_load(%s)""", [self.loadid])
        cursor.execute(
            """
            INSERT INTO load_errors (type, source, error, loadid, nodegroupid)
            SELECT 'tile', source_description, error_message, loadid, nodegroupid
            FROM load_staging
            WHERE loadid = %s AND passes_validation = false AND error_message IS NOT null
            """,
            [self.loadid],
        )
        return {"name": worksheet.title, "rows": row_count}

    def stage_excel_file(self, file, summary, cursor):
        if file.endswith("xlsx"):
            summary["files"][file]["worksheets"] = []
            workbook = load_workbook(filename=default_storage.open(os.path.join("uploadedfiles", "tmp", self.loadid, file)))
            try:
                nodegroup_id = workbook.active.cell(2, workbook.active.max_column).value
                graphid = str(Node.objects.filter(nodegroup_id=nodegroup_id)[:1].values_list('graph_id', flat=True)[0])
            except KeyError:
                cursor.execute(
                    """UPDATE load_event SET status = %s, load_end_time = %s WHERE loadid = %s""",
                    ("failed", datetime.now(), self.loadid),
                )
                raise ValueError(_("A graphid is not available in the metadata worksheet"))
            nodegroup_lookup, nodes = self.get_graph_tree(graphid)
            node_lookup = self.get_node_lookup(nodes)
            for worksheet in workbook.worksheets:
                if worksheet.title.lower() != "metadata":
                    details = self.process_worksheet(worksheet, cursor, node_lookup, nodegroup_lookup)
                    summary["files"][file]["worksheets"].append(details)
            cursor.execute(
                """UPDATE load_event SET load_details = %s WHERE loadid = %s""",
                (json.dumps(summary), self.loadid),
            )